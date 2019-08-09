'''Parse files in choosen folder and create spreadsheet containing
cutting tool usage data and file count by programmer.
Note - Toolist file location is hard coded.'''

# ToDo
# get rid of times used column
# format column widths

import os
import re
from collections import Counter
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog  # noqa: F401
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle


class Tool_list_Generator(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title('Tool List Generator')
        self.config(bg='light blue')

        # --------------------Combo Box

        label1 = tk.Label(self, text='Choose Machine')
        label1.grid(column=0, row=0)

        self.choose_machine_combo = ttk.Combobox(self, width=19)
        self.choose_machine_combo.set('MC12')
        self.choose_machine_combo['values'] = ('MC12', 'MC16', 'MH13', 'MH06')
        self.choose_machine_combo.bind("<<>ComboboxSelected>")
        self.choose_machine_combo.grid(column=1, row=0)

        # --------------------Listbox

        self.file_listbox = tk.Listbox(self, bg='light blue', width=80)
        self.file_listbox.grid(column=0, row=4, columnspan=5, sticky=tk.E+tk.W)

        # --------------------Buttons

        self.folder_pick = tk.Button(self, text="Select Folder",
                                     relief=tk.RAISED,
                                     width=16, bd=2, padx=10, pady=6)
        self.folder_pick.bind('<ButtonRelease-1>', self.choose_folder)
        self.folder_pick.grid(column=1, row=1)

        self.process = tk.Button(self, text="Process Data", relief=tk.RAISED,
                                 width=16, bd=2, padx=10, pady=6)
        self.process.grid(column=1, row=2)

        # --------------------Menu Bar
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        sub_menu = tk.Menu(menubar, tearoff=False)
        menubar.add_cascade(label='File', menu=sub_menu)
        # sub_menu.add_command(label='Open', command=open_file)

    def choose_folder(self, event):
        '''Opens folder selection dialog'''
        self.folder_selected = tk.filedialog.askdirectory()
        self.file_listbox.insert(tk.END, self.folder_selected)
        self.folder_pick.unbind('<ButtonRelease-1>')
        self.process.bind('<ButtonRelease-1>', self.extract)

    def extract(self, event):
        '''Creates dictionary of file names and tools in each'''
        os.chdir(self.folder_selected)
        files = os.listdir()
        pattern1 = re.compile(r'T\d+')
        pattern2 = re.compile(r'411Z91\d+-\w.*')
        match1 = filter(pattern2.search, files)
        self.target_files = []
        for item in match1:
            if os.path.isdir(item):
                continue
            bin_file = self.is_binary(item)
            if bin_file:
                continue
            self.target_files.append(item)
        self.result_dict = {}
        for item in self.target_files:
            with open(item, 'r') as f:
                file_contents = f.read()
                match2 = pattern1.findall(file_contents)
                self.result_dict[item] = set(match2)
            self.usage_count(self.result_dict)
        self.write_to_spreadsheet()

    def is_binary(self, file_name):
        ''' Fuction tries to open file as text and returns boolean'''
        try:
            with open(file_name, 'tr') as check_file:
                check_file.read()
                return False
        except UnicodeDecodeError:
            return True

    def usage_count(self, parsed_data):
        ''' Returns dictionary with tool number as key and
        number of times used as value'''
        all_file_tool_list = []
        for key, value in parsed_data.items():
            for item in value:
                all_file_tool_list.append(item)
        x = (s.strip('T') for s in all_file_tool_list)
        y = (s.replace('T', '') for s in x)
        new_tool_list = []
        for item in y:
            if item == '0' or item == '239' or int(item) > 300:
                continue
            new_tool_list.append(int(item))
        new_tool_list.sort()
        self.new_dict = {}
        temp_dict = Counter(new_tool_list)
        for k, v in temp_dict.items():
            self.new_dict[k] = v
        self.single_use(self.new_dict)

    def single_use(self, new_dict):
        '''Returns dictionary with tool number as key and
        file name as value for tools used in only one file'''
        single = []
        for keys, values in self.new_dict.items():
            if values == 1:
                single.append(keys)
        self.single_list = {}
        for tnum in single:
            for k, v in self.result_dict.items():
                for i in v:
                    if i == r'T{}'.format(tnum):
                        self.single_list[tnum] = k

    def get_ct_number(self, in_data):
        '''Gets CT number, description, holder and holder data
        from master tool list file.'''
        self.machine = self.choose_machine_combo.get()
        tl = 'C:/Users/dkjje/Desktop/Programming/Python_Projects/SS_Master_Tool_List/King Machine Cutting Tool List.xlsx'  # noqa: E501
        # tl = 'C:/Users/djennings/Documents/Programming/Python/SS_Master_Tool_List/King Machine Cutting Tool List.xlsx'  # noqa: E501
        wb = load_workbook(filename=tl)
        sh1 = wb.active
        t_data = {}
        for key in in_data:
            for cell1, cell2, cell3, cell4, cell5 in zip(sh1['A'], sh1['E'],
                                                         sh1['C'], sh1['J'],
                                                         sh1['Y']):
                if cell1.value == key and cell2.value == self.machine:
                    t_data[key] = (cell3.value, cell4.value, cell5.value)
        return t_data

    def extract_programmer(self):
        '''Retrieves programmer stats from files in folder'''
        dave_count = 0
        john_count = 0
        no_name = 0
        dave_list = []
        john_list = []
        unknown = []
        self.dave_list_set = set()
        self.john_list_set = set()
        self.unknown = set()
        programmer = {}
        for item in self.target_files:
            with open(item, 'r') as f:
                file_data = f.read()
                if 'DAVE' in file_data:
                    programmer[item] = 'Dave'
                    dave_list.append(item)
                    for item in dave_list:
                        item = item.replace('A', '').replace('B', '')
                        self.dave_list_set.add(item)
                    dave_count += 1
                elif 'JOHN' in file_data:
                    programmer[item] = 'John'
                    john_list.append(item)
                    for item in john_list:
                        item = item.replace('A', '').replace('B', '')
                        self.john_list_set.add(item)
                    john_count += 1
                else:
                    unknown.append(item)
                    programmer[item] = 'Unknown'
                    no_name += 1
        dave_percent = (dave_count / (dave_count + john_count + no_name) * 100)
        john_percent = (john_count / (dave_count + john_count + no_name) * 100)
        no_name_percent = (no_name / (dave_count + john_count + no_name) * 100)
        return (dave_percent, john_percent, no_name_percent,
                dave_count, john_count, no_name, dave_list, john_list, unknown)

    def max_length(self, eval_string):
        '''Function takes dictionary and returns length of longest
           key or value'''
        string_length = 0
        for keys, values in eval_string.items():
            if len(str(keys)) > len(str(values)) and (len(str(keys)) >
                                                      string_length):
                string_length = len(str(keys))
            elif len(str(values)) > string_length:
                string_length = len(str(values))
        return(string_length)

    def write_to_spreadsheet(self):
        '''Formats and writes data to spreadsheet.'''
        wb = Workbook()
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=False, size=11)
        bd = Side(style='thin', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        wb.add_named_style(highlight)  # Register named style
        sh1 = wb.active
        sh1.title = 'Tool Usage Frequency'
        sh1.append(['Tool Number', 'Times Used', 'CT Number', 'Description',
                    'Holder'])
        sh1['A1'].font = Font(bold=True, size=11)
        sh1['A1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh1['A1'].alignment = Alignment(horizontal='center')
        sh1['B1'].font = Font(bold=True, size=11)
        sh1['B1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh1['B1'].alignment = Alignment(horizontal='center')
        sh1['C1'].font = Font(bold=True, size=11)
        sh1['C1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh1['C1'].alignment = Alignment(horizontal='center')
        sh1['D1'].font = Font(bold=True, size=11)
        sh1['D1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh1['D1'].alignment = Alignment(horizontal='center')
        sh1['E1'].font = Font(bold=True, size=11)
        sh1['E1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh1['E1'].alignment = Alignment(horizontal='center')
        rnum = 2
        sh1.column_dimensions['A'].width = (11.95)
        sh1.column_dimensions['B'].width = (10.6)
        sh1.column_dimensions['C'].width = (10)
        col_width2 = 0
        sh1_ct_data = self.get_ct_number(self.new_dict)
        for keys, values in self.new_dict.items():  # keys is tool number, value is times used  # noqa: E501
            sh1.cell(row=rnum, column=1).value = int(keys)
            sh1.cell(row=rnum, column=2).value = int(values)
            sh1.cell(row=rnum, column=1).style = 'highlight'
            sh1.cell(row=rnum, column=2).style = 'highlight'
            sh1_tool_list_data = sh1_ct_data[keys]
            sh1_ct_num, sh1_description, sh1_holder = sh1_tool_list_data
            sh1.cell(row=rnum, column=3).value = sh1_ct_num
            sh1.cell(row=rnum, column=3).style = 'highlight'
            sh1.cell(row=rnum, column=4).value = sh1_description
            sh1.cell(row=rnum, column=4).style = 'highlight'
            sh1.cell(row=rnum, column=5).value = sh1_holder
            sh1.cell(row=rnum, column=5).style = 'highlight'
            if len(str(sh1_description)) > col_width2:
                col_width2 = len(str(sh1_description))
            rnum += 1
        sh1.column_dimensions['D'].width = (col_width2 * 1.125)
        # ----------------------------------------------------------
        sh2 = wb.create_sheet(title='Single Use List')
        wb.active = 2
        sh2.append(['Tool Number', 'Program Number', 'CT Number',
                   'Description'])
        sh2['A1'].font = Font(bold=True)
        sh2['A1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh2['A1'].alignment = Alignment(horizontal='center')
        sh2['B1'].font = Font(bold=True)
        sh2['B1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh2['B1'].alignment = Alignment(horizontal='center')
        sh2['C1'].font = Font(bold=True, size=11)
        sh2['C1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh2['C1'].alignment = Alignment(horizontal='center')
        sh2['D1'].font = Font(bold=True, size=11)
        sh2['D1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh2['D1'].alignment = Alignment(horizontal='center')
        rnum = 2
        col_width = (self.max_length(self.single_list))
        sh2.column_dimensions['A'].width = (11.95)
        sh2.column_dimensions['B'].width = (col_width * 1.125)
        sh2.column_dimensions['C'].width = (10)
        col_width4 = 0
        col_width5 = 0
        for keys, values in self.single_list.items():
            sh2_tool_list_data = sh1_ct_data[keys]
            sh2_ct_num, sh2_description, holder = sh2_tool_list_data
            sh2.cell(row=rnum, column=1).value = int(keys)
            sh2.cell(row=rnum, column=2).value = (values)
            sh2.cell(row=rnum, column=1).style = 'highlight'
            sh2.cell(row=rnum, column=2).style = 'highlight'
            sh2.cell(row=rnum, column=3).value = sh2_ct_num
            sh2.cell(row=rnum, column=3).style = 'highlight'
            sh2.cell(row=rnum, column=4).value = sh2_description
            sh2.cell(row=rnum, column=4).style = 'highlight'
            if len(str(sh2_description)) > col_width4:
                col_width4 = len(str(sh2_description))
            if len(str(holder)) > col_width5:
                col_width5 = len(str(holder))
            rnum += 1
        sh2.column_dimensions['D'].width = (col_width4 * 1.125)
        sh2.column_dimensions['E'].width = (col_width5 * 1.125)

        # ----------------------------------------------------------
        sh3 = wb.create_sheet(title='Programmer')
        wb.active = 3
        programmed_by = self.extract_programmer()
        dp, jp, np, dc, jc, nc, dl, jl, ul = programmed_by
        sh3.append(['Part Number', 'Programmer'])
        sh3['A1'].font = Font(bold=True, size=11)
        sh3['A1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh3['A1'].alignment = Alignment(horizontal='center')
        sh3['B1'].font = Font(bold=True)
        sh3['B1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        sh3['B1'].alignment = Alignment(horizontal='center')
        rnum = 2
        for item in self.dave_list_set:
            sh3.cell(row=rnum, column=2).value = 'Dave'
            sh3.cell(row=2, column=2).style = 'highlight'
            sh3.cell(row=rnum, column=1).value = item
            sh3.cell(row=2, column=1).style = 'highlight'
            rnum += 1
        for item in self.john_list_set:
            sh3.cell(row=rnum, column=2).value = 'John'
            sh3.cell(row=2, column=2).style = 'highlight'
            sh3.cell(row=rnum, column=1).value = item
            sh3.cell(row=2, column=1).style = 'highlight'
            rnum += 1
        for item in self.unknown:
            sh3.cell(row=rnum, column=2).value = 'Unknown'
            sh3.cell(row=2, column=2).style = 'highlight'
            sh3.cell(row=rnum, column=1).value = item
            sh3.cell(row=2, column=1).style = 'highlight'
            rnum += 1
        save_name = (('{}/{} Tool Usage Data.xlsx').format
                     (self.folder_selected, self.machine))
        wb.save(save_name)
        self.file_listbox.insert(tk.END, 'Operation Complete')
        self.file_listbox.see(tk.END)
        os.startfile(save_name)


root = Tool_list_Generator()
root.mainloop()
