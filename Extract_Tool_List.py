''' Generate list of tool numbers and usage frequency from specified directory
and output to new .xlsx sheet. Generate list of tools used in only one file
and output file name and tool number to new sheet. '''

import re
import os
from collections import Counter
import tkinter as tk
# from tkinter import ttk
from tkinter import filedialog  # noqa: F401
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle


root = tk.Tk()
root.title('Extract Tool List From Machine Library')
root.folder_selected = ''


def choose_folder(event):
    '''Opens folder selection dialog'''
    root.folder_selected = tk.filedialog.askdirectory()
    file_listbox.insert(tk.END, root.folder_selected)
    folder_pick.unbind('<ButtonRelease-1>')
    process.bind('<ButtonRelease-1>', extract)


def is_binary(file_name):
    try:
        with open(file_name, 'tr') as check_file:  # try open file in text mode
            check_file.read()
            return False
    except UnicodeDecodeError:  # if fail then file is non-text (binary)
        return True


def extract(event):
    '''Retrieves tool numbers used in all files in folder'''
    os.chdir(root.folder_selected)
    files = os.listdir()
    pattern1 = re.compile(r'T\d+')
    pattern2 = re.compile(r'411Z91\d+-\w.*')
    tool_list = []  # Combination of sets w/ duplicates - needed for Counter
    match = filter(pattern2.search, files)
    target_files = []
    for item in match:
        if os.path.isdir(item):
            continue
        bin_file = is_binary(item)
        if bin_file:
            continue
        target_files.append(item)
        with open(item, 'r') as f:
            get_t_number = f.read()
            match2 = pattern1.findall(get_t_number)
            tool_set = set(match2)
            for item in tool_set:
                tool_list.append(item)
    # print(tool_list)
    x = (s.strip('T') for s in tool_list)  # Need to understand generators
    y = (s.replace('T', '') for s in x)
    new_tool_list = []
    for item in y:
        if item == '0' or item == '239' or int(item) > 300:
            continue
        new_tool_list.append(int(item))
    new_tool_list.sort()
    new_dict = Counter(new_tool_list)
# --------------------------------------------------------
    single_list = []
    for keys, values in new_dict.items():
        if values == 1:
            single_list.append(keys)  # List of tool #'s used in only one file
    single_use_tool_dict = {}
    for item in target_files:
        if os.path.isdir(item):
            continue
        bin_file = is_binary(item)
        if bin_file:
            continue
        with open(item, 'r') as f:
            get_single_use_name = f.read()
        for tnum in single_list:
            pattern2 = re.compile(r'T({})\D'.format(tnum))
            match2 = re.search(pattern2, get_single_use_name)
            if match2:
                single_use_tool_dict[tnum] = item

    wb = Workbook()
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=False, size=11)
    bd = Side(style='thin', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(highlight)  # Register named style
    sh1 = wb.active
    sh1.title = 'Tool Usage Frequency'
    sh1.append(['Tool Number', 'Times Used'])
    sh1['A1'].font = Font(bold=True, size=11)
    sh1['A1'].alignment = Alignment(horizontal='center')
    sh1['B1'].font = Font(bold=True, size=11)
    sh1['B1'].alignment = Alignment(horizontal='center')
    rnum = 2
    sh1.column_dimensions['A'].width = (11.95)
    sh1.column_dimensions['B'].width = (10.6)
    for keys, values in new_dict.items():
        sh1.cell(row=rnum, column=1).value = int(keys)
        sh1.cell(row=rnum, column=2).value = int(values)

        sh1.cell(row=rnum, column=1).style = 'highlight'
        sh1.cell(row=rnum, column=2).style = 'highlight'

        rnum += 1
# ----------------------------------------------------------
    sh2 = wb.create_sheet(title='Single Use List')
    wb.active = 2
    sh2.append(['Tool Number', 'Program Number'])
    sh2['A1'].font = Font(bold=True)
    sh2['A1'].alignment = Alignment(horizontal='center')
    sh2['B1'].font = Font(bold=True)
    sh2['B1'].alignment = Alignment(horizontal='center')
    rnum = 2
    col_width = (max_length(single_use_tool_dict))
    sh2.column_dimensions['A'].width = (11.95)
    sh2.column_dimensions['B'].width = (col_width * 1.125)
    print('return = ', col_width)
    for keys, values in single_use_tool_dict.items():
        sh2.cell(row=rnum, column=1).value = int(keys)
        sh2.cell(row=rnum, column=2).value = (values)
        rnum += 1

    wb.save('C:/Users/djennings/Documents/Programming/Python/SS_Master_Tool_List/TestWrite.xlsx')
    file_listbox.insert(tk.END, 'Operation Complete')
    file_listbox.see(tk.END)


def max_length(eval_string):
    '''Function takes dictionary and returns length of longest key or value'''
    string_length = 0
    for keys, values in eval_string.items():
        if len(str(keys)) > len(str(values)) and (len(str(keys)) >
                                                  string_length):
            string_length = len(str(keys))
        elif len(str(values)) > string_length:
            string_length = len(str(values))
    return(string_length)


# --------------------Listbox


file_listbox = tk.Listbox(root, bg='light blue', width=80)
file_listbox.grid(column=0, row=3, columnspan=5, sticky=tk.E+tk.W)

# --------------------Buttons

folder_pick = tk.Button(root, text="Select Folder", relief=tk.RAISED,
                        width=16, bd=2, padx=10, pady=6)
folder_pick.bind('<ButtonRelease-1>', choose_folder)
folder_pick.grid(column=2, row=0)

process = tk.Button(root, text="Process Data", relief=tk.RAISED,
                    width=16, bd=2, padx=10, pady=6)
process.grid(column=2, row=1)

# --------------------Menu Bar
menubar = tk.Menu(root)
root.config(menu=menubar)
file_menu = tk.Menu(menubar, tearoff=False)
menubar.add_cascade(label='File', menu=file_menu)
# file_menu.add_command(label='Open', command=open_traceback)


root.mainloop()
