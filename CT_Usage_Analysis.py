'''Parse files in choosen folder and create spreadsheet containing
cutting tool usage data and file count by programmer.
Note - Toolist file location is hard coded.'''

# ToDo
# get base part number from dave_list and john_list
# unpack on line 276 has holder - see where it comes from
# get rid of times used column
# format column widths
# Wrap in class
# Add programmer list sheet

import os
import re
from collections import Counter
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog  # noqa: F401
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle


root = tk.Tk()
root.title('Extract Tool List From Machine Library')

# Class variables

root.new_dict = {}
root.single_list = {}
root.result_dict = {}
root.machine = ''
root.folder_selected = ''
root.target_files = []
root.programmer = {}


def choose_folder(event):
    '''Opens folder selection dialog'''
    root.folder_selected = tk.filedialog.askdirectory()
    file_listbox.insert(tk.END, root.folder_selected)
    folder_pick.unbind('<ButtonRelease-1>')
    process.bind('<ButtonRelease-1>', extract)


def extract(event):
    os.chdir(root.folder_selected)
    files = os.listdir()
    pattern1 = re.compile(r'T\d+')
    pattern2 = re.compile(r'(411Z91\d+-\w*\.(MC12|mc12))|(112A5251-60.*)')
    # pattern2 = re.compile(r'112A5251-60.*')

    match1 = filter(pattern2.search, files)
    for item in match1:
        if os.path.isdir(item):
            continue
        bin_file = is_binary(item)
        if bin_file:
            continue
        root.target_files.append(item)
    for item in root.target_files:
        with open(item, 'r') as f:
            file_contents = f.read()
            match2 = pattern1.findall(file_contents)
            root.result_dict[item] = set(match2)
        usage_count(root.result_dict)
    # print(root.target_files)
    write_to_spreadsheet()


def is_binary(file_name):
    ''' Fuction tries to open file as test and returns boolean'''
    try:
        with open(file_name, 'tr') as check_file:
            check_file.read()
            return False
    except UnicodeDecodeError:
        return True


def usage_count(parsed_data):
    ''' Returns dictionary with tool number as key and
    number of times used as value'''
    all_file_tool_list = []
    for key, value in parsed_data.items():
        for item in value:
            all_file_tool_list.append(item)
    x = (s.strip('T') for s in all_file_tool_list)
    y = (s.replace('T', '') for s in x)
    new_tool_list = []
    for item in y:
        if item == '0' or item == '239' or int(item) > 300:
            continue
        new_tool_list.append(int(item))
    new_tool_list.sort()
    temp_dict = Counter(new_tool_list)
    for k, v in temp_dict.items():
        root.new_dict[k] = v
    print(root.new_dict)
    single_use(root.new_dict)


def single_use(new_dict):
    '''Returns dictionary with tool number as key and
    file name as value for tools used in only one file'''
    single = []
    for keys, values in root.new_dict.items():
        if values == 1:
            single.append(keys)
    for tnum in single:
        for k, v in root.result_dict.items():
            for i in v:
                if i == r'T{}'.format(tnum):
                    root.single_list[tnum] = k


def get_ct_number(in_data):
    '''Gets CT number from master tool list file.'''
    root.machine = choose_machine_combo.get()
    # tl = 'C:/Users/dkjje/Desktop/Programming/Python_Projects/SS_Master_Tool_List/King Machine Cutting Tool List.xlsx'
    tl = 'C:/Users/djennings/Documents/Programming/Python/SS_Master_Tool_List/King Machine Cutting Tool List.xlsx'
    wb = load_workbook(filename=tl)
    sh1 = wb.active
    t_data = {}
    for key in in_data:
        for cell1, cell2, cell3, cell4, cell5 in zip(sh1['A'], sh1['E'],
                                                     sh1['C'], sh1['J'],
                                                     sh1['Y']):
            if cell1.value == key and cell2.value == root.machine:
                t_data[key] = (cell3.value, cell4.value, cell5.value)
    return t_data


def extract_programmer():
    '''Retrieves programmer stats from files in folder'''
    dave_count = 0
    john_count = 0
    no_name = 0
    dave_list = []
    john_list = []
    unknown = []
    root.dave_list_set = set()
    root.john_list_set = set()
    root.unknown_list_set = set()

    for item in root.target_files:
        with open(item, 'r') as f:
            file_data = f.read()
            if 'DAVE' in file_data:
                root.programmer[item] = 'Dave'
                dave_list.append(item)
                for item in dave_list:
                    item = item.replace('A.', '.').replace('B.', '.')
                    root.dave_list_set.add(item)
                dave_count += 1
            elif 'JOHN' in file_data:
                root.programmer[item] = 'John'
                john_list.append(item)
                for item in john_list:
                    item = item.replace('A.', '.').replace('B.', '.')
                    root.john_list_set.add(item)
                john_count += 1
            else:
                unknown.append(item)
                root.programmer[item] = 'Unknown'
                for item in unknown:
                    item = item.replace('A.', '.').replace('B.', '.')
                    root.unknown_list_set.add(item)
                no_name += 1
    root.sorted_dave = sorted(root.dave_list_set)
    root.sorted_john = sorted(root.john_list_set)
    # print(root.sorted_dave)
    dave_percent = (dave_count / (dave_count + john_count + no_name) * 100)
    john_percent = (john_count / (dave_count + john_count + no_name) * 100)
    no_name_percent = (no_name / (dave_count + john_count + no_name) * 100)
    # print(root.programmer)
    return (dave_percent, john_percent, no_name_percent,
            dave_count, john_count, no_name, dave_list, john_list, unknown)


def max_length(eval_string):
    '''Function takes dictionary and returns length of longest key or value'''
    string_length = 0
    for keys, values in eval_string.items():
        if len(str(keys)) > len(str(values)) and (len(str(keys)) >
                                                  string_length):
            string_length = len(str(keys))
        elif len(str(values)) > string_length:
            string_length = len(str(values))
    return(string_length)

# --------------------Combo Box


label1 = tk.Label(root, text='Choose Machine')
label1.grid(column=0, row=0)

choose_machine_combo = ttk.Combobox(root, width=19)
choose_machine_combo.set('MC12')
choose_machine_combo['values'] = ('MC12', 'MH13', 'MH06')
choose_machine_combo.bind("<<>ComboboxSelected>")
choose_machine_combo.grid(column=1, row=0)

# --------------------Listbox

file_listbox = tk.Listbox(root, bg='light blue', width=80)
file_listbox.grid(column=0, row=4, columnspan=5, sticky=tk.E+tk.W)

# --------------------Buttons

folder_pick = tk.Button(root, text="Select Folder", relief=tk.RAISED,
                        width=16, bd=2, padx=10, pady=6)
folder_pick.bind('<ButtonRelease-1>', choose_folder)
folder_pick.grid(column=1, row=1)

process = tk.Button(root, text="Process Data", relief=tk.RAISED,
                    width=16, bd=2, padx=10, pady=6)
process.grid(column=1, row=2)

# --------------------Menu Bar
menubar = tk.Menu(root)
root.config(menu=menubar)
sub_menu = tk.Menu(menubar, tearoff=False)
menubar.add_cascade(label='File', menu=sub_menu)
# sub_menu.add_command(label='Open', command=open_file)


def write_to_spreadsheet():
    '''Formats and writes data to spreadsheet.'''
    wb = Workbook()
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=False, size=11)
    bd = Side(style='thin', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(highlight)  # Register named style
    sh1 = wb.active
    sh1.title = 'Tool Usage Frequency'
    sh1.append(['Tool Number', 'CT Number', 'Description',
                'Holder'])
    sh1['A1'].font = Font(bold=True, size=11)
    sh1['A1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sh1['A1'].alignment = Alignment(horizontal='center')
    sh1['B1'].font = Font(bold=True, size=11)
    sh1['B1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sh1['B1'].alignment = Alignment(horizontal='center')
    sh1['C1'].font = Font(bold=True, size=11)
    sh1['C1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sh1['C1'].alignment = Alignment(horizontal='center')
    sh1['D1'].font = Font(bold=True, size=11)
    sh1['D1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sh1['D1'].alignment = Alignment(horizontal='center')
    rnum = 2
    sh1.column_dimensions['A'].width = (11.95)
    sh1.column_dimensions['B'].width = (10.6)
    sh1.column_dimensions['C'].width = (10)
    col_width2 = 0
    sh1_ct_data = get_ct_number(root.new_dict)
    for keys, values in root.new_dict.items():
        sh1_tool_list_data = sh1_ct_data[keys]  # keys is tool number, value is times used
        sh1.cell(row=rnum, column=1).value = int(keys)
        # sh1.cell(row=rnum, column=2).value = int(values)
        sh1.cell(row=rnum, column=1).style = 'highlight'
        # sh1.cell(row=rnum, column=2).style = 'highlight'
        sh1_ct_num, sh1_description, sh1_holder = sh1_tool_list_data
        sh1.cell(row=rnum, column=2).value = sh1_ct_num
        sh1.cell(row=rnum, column=2).style = 'highlight'
        sh1.cell(row=rnum, column=3).value = sh1_description
        sh1.cell(row=rnum, column=3).style = 'highlight'
        sh1.cell(row=rnum, column=4).value = sh1_holder
        sh1.cell(row=rnum, column=4).style = 'highlight'
        if len(str(sh1_description)) > col_width2:
            col_width2 = len(str(sh1_description))
        rnum += 1
    sh1.column_dimensions['C'].width = (col_width2 * 1.125)
    # ----------------------------------------------------------
    # sh2 = wb.create_sheet(title='Single Use List')
    # wb.active = 2
    # sh2.append(['Tool Number', 'Program Number', 'CT Number', 'Description'])
    # sh2['A1'].font = Font(bold=True)
    # sh2['A1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    # sh2['A1'].alignment = Alignment(horizontal='center')
    # sh2['B1'].font = Font(bold=True)
    # sh2['B1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    # sh2['B1'].alignment = Alignment(horizontal='center')
    # sh2['C1'].font = Font(bold=True, size=11)
    # sh2['C1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    # sh2['C1'].alignment = Alignment(horizontal='center')
    # sh2['D1'].font = Font(bold=True, size=11)
    # sh2['D1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    # sh2['D1'].alignment = Alignment(horizontal='center')
    # rnum = 2
    # col_width = (max_length(root.single_list))
    # sh2.column_dimensions['A'].width = (11.95)
    # sh2.column_dimensions['B'].width = (col_width * 1.125)
    # sh2.column_dimensions['C'].width = (10)
    # col_width3 = 0
    # for keys, values in root.single_list.items():
    #     sh2_tool_list_data = sh1_ct_data[keys]
    #     sh2_ct_num, sh2_description, holder = sh2_tool_list_data
    #     sh2.cell(row=rnum, column=1).value = int(keys)
    #     sh2.cell(row=rnum, column=2).value = (values)
    #     sh2.cell(row=rnum, column=1).style = 'highlight'
    #     sh2.cell(row=rnum, column=2).style = 'highlight'
    #     sh2.cell(row=rnum, column=3).value = sh2_ct_num
    #     sh2.cell(row=rnum, column=3).style = 'highlight'
    #     sh2.cell(row=rnum, column=4).value = sh2_description
    #     sh2.cell(row=rnum, column=4).style = 'highlight'
    #     if len(str(sh2_description)) > col_width3:
    #         col_width3 = len(str(sh2_description))
    #     rnum += 1
    # sh2.column_dimensions['D'].width = (col_width3 * 1.125)

    # ----------------------------------------------------------
    sh2 = wb.create_sheet(title='Programmer')
    wb.active = 2
    programmed_by = extract_programmer()
    dp, jp, np, dc, jc, nc, dl, jl, ul = programmed_by
    sh2.append(['Part Number', 'Programmer'])
    sh2['A1'].font = Font(bold=True, size=11)
    sh2['A1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sh2['A1'].alignment = Alignment(horizontal='center')
    sh2['B1'].font = Font(bold=True)
    sh2['B1'].border = Border(left=bd, top=bd, right=bd, bottom=bd)
    sh2['B1'].alignment = Alignment(horizontal='center')
    rnum = 2
    for item in root.sorted_dave:

        sh2.cell(row=rnum, column=2).value = 'Dave'
        sh2.cell(row=2, column=2).style = 'highlight'
        sh2.cell(row=rnum, column=1).value = item
        sh2.cell(row=2, column=1).style = 'highlight'
        rnum += 1

    for item in root.sorted_john:

        sh2.cell(row=rnum, column=2).value = 'John'
        sh2.cell(row=2, column=2).style = 'highlight'
        sh2.cell(row=rnum, column=1).value = item
        sh2.cell(row=2, column=1).style = 'highlight'
        rnum += 1

    for item in root.unknown_list_set:

        sh2.cell(row=rnum, column=2).value = 'Unknown'
        sh2.cell(row=2, column=2).style = 'highlight'
        sh2.cell(row=rnum, column=1).value = item
        sh2.cell(row=2, column=1).style = 'highlight'
        rnum += 1

    save_name = (('{}/{} Tool Usage Data.xlsx').format
                 (root.folder_selected, root.machine))
    wb.save(save_name)
    file_listbox.insert(tk.END, 'Operation Complete')
    file_listbox.see(tk.END)
    os.startfile(save_name)


root.mainloop()
